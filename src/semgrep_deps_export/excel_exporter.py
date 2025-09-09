"""
Excel export functionality for Semgrep dependencies data.

Creates formatted XLSX files with Dependencies and Vulnerabilities sheets.
"""

import logging
import os
from datetime import datetime
from typing import List, Optional, Dict, Any
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

from .data_processor import ProcessedDependency, ProcessedVulnerability
from .config import Config


logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports processed dependency data to Excel format."""
    
    def __init__(self, config: Config):
        self.config = config
        self.workbook = Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        self.cell_alignment = Alignment(horizontal="left", vertical="center")
        self.number_alignment = Alignment(horizontal="right", vertical="center")
        
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Severity colors
        self.severity_colors = {
            "Critical": "FF0000",  # Red
            "High": "FF6600",      # Orange
            "Medium": "FFCC00",    # Yellow
            "Low": "99CC00",       # Light Green
            "Info": "CCCCCC"       # Gray
        }
    
    def _generate_filename(self) -> str:
        """Generate filename based on deployment ID and timestamp."""
        if self.config.output_path:
            return self.config.output_path
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"semgrep_dependencies_{self.config.deployment_id}_{timestamp}.xlsx"
        
        # Use output directory if specified, otherwise use 'output' directory
        if self.config.output_dir:
            output_dir = self.config.output_dir
        else:
            output_dir = os.path.join(os.getcwd(), "output")
        
        return os.path.join(output_dir, filename)
    
    def _generate_filtered_filename(self) -> str:
        """Generate filename for filtered dependencies with bad/review licenses."""        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"bad_review_license_semgrep_dependencies_{self.config.deployment_id}_{timestamp}.xlsx"
        
        # Use output directory if specified, otherwise use 'output' directory
        if self.config.output_dir:
            output_dir = self.config.output_dir
        else:
            output_dir = os.path.join(os.getcwd(), "output")
        
        return os.path.join(output_dir, filename)
    
    def _generate_policy_blocked_filename(self) -> str:
        """Generate filename for LICENSE_POLICY_SETTING_BLOCK dependencies."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"policy_blocked_semgrep_dependencies_{self.config.deployment_id}_{timestamp}.xlsx"
        
        # Use output directory if specified, otherwise use 'output' directory
        if self.config.output_dir:
            output_dir = self.config.output_dir
        else:
            output_dir = os.path.join(os.getcwd(), "output")
        
        return os.path.join(output_dir, filename)
    
    def _generate_policy_comment_filename(self) -> str:
        """Generate filename for LICENSE_POLICY_SETTING_COMMENT dependencies."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"policy_comment_semgrep_dependencies_{self.config.deployment_id}_{timestamp}.xlsx"
        
        # Use output directory if specified, otherwise use 'output' directory
        if self.config.output_dir:
            output_dir = self.config.output_dir
        else:
            output_dir = os.path.join(os.getcwd(), "output")
        
        return os.path.join(output_dir, filename)
    
    def _generate_ecosystem_pypi_filename(self) -> str:
        """Generate filename for PyPI ecosystem dependencies."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ecosystem_pypi_semgrep_dependencies_{self.config.deployment_id}_{timestamp}.xlsx"
        
        # Use output directory if specified, otherwise use 'output' directory
        if self.config.output_dir:
            output_dir = self.config.output_dir
        else:
            output_dir = os.path.join(os.getcwd(), "output")
        
        return os.path.join(output_dir, filename)
    
    def _create_dependencies_sheet(self, dependencies: List[ProcessedDependency], include_license_columns: bool = True, apply_license_coloring: bool = True) -> Worksheet:
        """Create the Dependencies worksheet."""
        logger.info("Creating Dependencies sheet...")
        
        ws = self.workbook.create_sheet("Dependencies")
        
        # Define headers based on whether to include license columns
        if include_license_columns:
            headers = [
                "Repository Name",
                "Name",
                "Version",
                "Ecosystem",
                "Package Manager",
                "Transitivity",
                "Bad_License",
                "Review_License",
                "Licenses"
            ]
        else:
            headers = [
                "Repository Name",
                "Name",
                "Version",
                "Ecosystem",
                "Package Manager",
                "Transitivity",
                "Licenses"
            ]
        
        # Set headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Add data
        for row, dep in enumerate(dependencies, 2):
            ws.cell(row=row, column=1, value=dep.repository_name)
            ws.cell(row=row, column=2, value=dep.name)
            ws.cell(row=row, column=3, value=dep.version)
            ws.cell(row=row, column=4, value=dep.ecosystem)
            ws.cell(row=row, column=5, value=dep.package_manager)
            ws.cell(row=row, column=6, value=dep.transitivity)
            
            if include_license_columns:
                ws.cell(row=row, column=7, value=dep.bad_license)
                ws.cell(row=row, column=8, value=dep.review_license)
                ws.cell(row=row, column=9, value=dep.licenses)
            else:
                ws.cell(row=row, column=7, value=dep.licenses)
            
            # Apply styles to data cells
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                
                # All columns use left alignment now (no number columns remaining)
                cell.alignment = self.cell_alignment
            
            # Apply highlighting for license types only if enabled and license columns are included
            if apply_license_coloring and include_license_columns:
                if dep.bad_license and dep.review_license:
                    # Both bad and review licenses - use combined formatting
                    self._apply_dual_license_formatting(ws, row, len(headers))
                elif dep.bad_license:
                    # Bad license only - red highlighting
                    self._apply_bad_license_formatting(ws, row, len(headers))
                elif dep.review_license:
                    # Review license only - yellow highlighting
                    self._apply_review_license_formatting(ws, row, len(headers))
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = len(headers[col-1])
            
            for row in range(2, len(dependencies) + 2):
                try:
                    cell_value = str(ws.cell(row=row, column=col).value)
                    max_length = max(max_length, len(cell_value))
                except:
                    pass
            
            # Set column width with some padding
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        logger.info(f"Dependencies sheet created with {len(dependencies)} rows")
        return ws
    
    def _apply_bad_license_formatting(self, ws: Worksheet, row: int, num_cols: int) -> None:
        """Apply red background formatting to bad license rows."""
        bad_license_fill = PatternFill(
            start_color="FFCCCC",  # Light red background
            end_color="FFCCCC",
            fill_type="solid"
        )
        
        # Apply to all cells in the row
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            # Preserve existing border formatting
            if cell.border:
                cell.fill = bad_license_fill
            else:
                cell.fill = bad_license_fill
                cell.border = self.border
    
    def _apply_review_license_formatting(self, ws: Worksheet, row: int, num_cols: int) -> None:
        """Apply yellow background formatting to review license rows."""
        review_license_fill = PatternFill(
            start_color="FFFFCC",  # Light yellow background
            end_color="FFFFCC",
            fill_type="solid"
        )
        
        # Apply to all cells in the row
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            # Preserve existing border formatting
            if cell.border:
                cell.fill = review_license_fill
            else:
                cell.fill = review_license_fill
                cell.border = self.border
    
    def _apply_dual_license_formatting(self, ws: Worksheet, row: int, num_cols: int) -> None:
        """Apply combined formatting for dependencies with both bad and review licenses."""
        dual_license_fill = PatternFill(
            start_color="FFDDAA",  # Light orange background (mix of red and yellow)
            end_color="FFDDAA",
            fill_type="solid"
        )
        
        # Apply to all cells in the row
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            # Preserve existing border formatting
            if cell.border:
                cell.fill = dual_license_fill
            else:
                cell.fill = dual_license_fill
                cell.border = self.border
    
    def _create_vulnerabilities_sheet(self, vulnerabilities: List[ProcessedVulnerability]) -> Optional[Worksheet]:
        """Create the Vulnerabilities worksheet."""
        if not vulnerabilities:
            logger.info("No vulnerabilities to export, skipping Vulnerabilities sheet")
            return None
            
        logger.info("Creating Vulnerabilities sheet...")
        
        ws = self.workbook.create_sheet("Vulnerabilities")
        
        # Define headers
        headers = [
            "Dependency Name",
            "Dependency Version",
            "Vulnerability ID",
            "Severity",
            "Description"
        ]
        
        # Set headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Add data
        for row, vuln in enumerate(vulnerabilities, 2):
            ws.cell(row=row, column=1, value=vuln.dependency_name)
            ws.cell(row=row, column=2, value=vuln.dependency_version)
            ws.cell(row=row, column=3, value=vuln.vulnerability_id)
            ws.cell(row=row, column=4, value=vuln.severity)
            ws.cell(row=row, column=5, value=vuln.description)
            
            # Apply styles to data cells
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = self.cell_alignment
                
                # Color-code severity
                if col == 4:  # Severity column
                    severity = vuln.severity
                    if severity in self.severity_colors:
                        cell.fill = PatternFill(
                            start_color=self.severity_colors[severity],
                            end_color=self.severity_colors[severity],
                            fill_type="solid"
                        )
                        # Use white text for dark backgrounds
                        if severity in ["Critical", "High"]:
                            cell.font = Font(color="FFFFFF", bold=True)
                        else:
                            cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = len(headers[col-1])
            
            for row in range(2, min(len(vulnerabilities) + 2, 100)):  # Check first 100 rows for performance
                try:
                    cell_value = str(ws.cell(row=row, column=col).value)
                    max_length = max(max_length, len(cell_value))
                except:
                    pass
            
            # Set column width with some padding
            if col == 5:  # Description column - limit width
                ws.column_dimensions[column_letter].width = min(max_length + 2, 80)
            else:
                ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        logger.info(f"Vulnerabilities sheet created with {len(vulnerabilities)} rows")
        return ws
    
    def _create_summary_sheet(self, summary: Dict[str, Any]) -> Worksheet:
        """Create a summary sheet with processing statistics."""
        logger.info("Creating Summary sheet...")
        
        ws = self.workbook.create_sheet("Summary", 0)  # Insert as first sheet
        
        # Title
        ws.cell(row=1, column=1, value="Semgrep Dependencies Export Summary")
        ws.cell(row=1, column=1).font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")
        
        # Export metadata
        ws.cell(row=3, column=1, value="Export Details")
        ws.cell(row=3, column=1).font = Font(bold=True)
        
        ws.cell(row=4, column=1, value="Deployment ID:")
        ws.cell(row=4, column=2, value=self.config.deployment_id)
        
        ws.cell(row=5, column=1, value="Export Date:")
        ws.cell(row=5, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC"))
        
        # Dependencies summary
        ws.cell(row=7, column=1, value="Dependencies Summary")
        ws.cell(row=7, column=1).font = Font(bold=True)
        
        ws.cell(row=8, column=1, value="Total Dependencies:")
        ws.cell(row=8, column=2, value=summary["dependencies"]["total"])
        
        ws.cell(row=9, column=1, value="With Vulnerabilities:")
        ws.cell(row=9, column=2, value=summary["dependencies"]["with_vulnerabilities"])
        
        ws.cell(row=10, column=1, value="Without Vulnerabilities:")
        ws.cell(row=10, column=2, value=summary["dependencies"]["without_vulnerabilities"])
        
        # Vulnerabilities summary
        ws.cell(row=12, column=1, value="Vulnerabilities Summary")
        ws.cell(row=12, column=1).font = Font(bold=True)
        
        ws.cell(row=13, column=1, value="Total Vulnerabilities:")
        ws.cell(row=13, column=2, value=summary["vulnerabilities"]["total"])
        
        ws.cell(row=14, column=1, value="Critical:")
        ws.cell(row=14, column=2, value=summary["vulnerabilities"]["critical"])
        
        ws.cell(row=15, column=1, value="High:")
        ws.cell(row=15, column=2, value=summary["vulnerabilities"]["high"])
        
        ws.cell(row=16, column=1, value="Medium:")
        ws.cell(row=16, column=2, value=summary["vulnerabilities"]["medium"])
        
        ws.cell(row=17, column=1, value="Low:")
        ws.cell(row=17, column=2, value=summary["vulnerabilities"]["low"])
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        return ws
    
    def export(self, dependencies: List[ProcessedDependency], vulnerabilities: List[ProcessedVulnerability], summary: Dict[str, Any]) -> str:
        """Export data to Excel file."""
        output_path = self._generate_filename()
        
        logger.info(f"Starting Excel export to {output_path}")
        
        try:
            # Create sheets
            self._create_dependencies_sheet(dependencies)
            
            if vulnerabilities:
                self._create_vulnerabilities_sheet(vulnerabilities)
            
            # Ensure output directory exists
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            
            # Save workbook
            self.workbook.save(output_path)
            
            # Get file size for logging
            file_size = os.path.getsize(output_path)
            file_size_mb = file_size / (1024 * 1024)
            
            logger.info(f"Excel export completed successfully:")
            logger.info(f"  - File: {output_path}")
            logger.info(f"  - Size: {file_size_mb:.2f} MB")
            logger.info(f"  - Sheets: Dependencies" + (", Vulnerabilities" if vulnerabilities else ""))
            
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to export Excel file: {str(e)}")
            raise Exception(f"Excel export failed: {str(e)}")

    def export_filtered(self, dependencies: List[ProcessedDependency], vulnerabilities: List[ProcessedVulnerability], summary: Dict[str, Any]) -> Optional[str]:
        """Export filtered data (bad_license OR review_license == True) to separate Excel file."""
        # Filter dependencies to only those with bad or review licenses
        filtered_dependencies = [dep for dep in dependencies if dep.bad_license or dep.review_license]
        
        # If no problematic dependencies found, skip creating filtered file
        if not filtered_dependencies:
            logger.info("No dependencies with bad or review licenses found, skipping filtered export")
            return None
            
        # Filter vulnerabilities to only those associated with filtered dependencies
        filtered_dep_names = {f"{dep.name}:{dep.version}" for dep in filtered_dependencies}
        filtered_vulnerabilities = [
            vuln for vuln in vulnerabilities 
            if f"{vuln.dependency_name}:{vuln.dependency_version}" in filtered_dep_names
        ]
        
        # Create new workbook for filtered export
        filtered_workbook = Workbook()
        filtered_workbook.remove(filtered_workbook.active)  # Remove default sheet
        original_workbook = self.workbook  # Store original workbook
        
        try:
            # Temporarily use filtered workbook
            self.workbook = filtered_workbook
            
            output_path = self._generate_filtered_filename()
            logger.info(f"Starting filtered Excel export to {output_path}")
            logger.info(f"  - Filtered dependencies: {len(filtered_dependencies)} (from {len(dependencies)} total)")
            logger.info(f"  - Filtered vulnerabilities: {len(filtered_vulnerabilities)} (from {len(vulnerabilities)} total)")
            
            # Create sheets with filtered data
            self._create_dependencies_sheet(filtered_dependencies)
            
            if filtered_vulnerabilities:
                self._create_vulnerabilities_sheet(filtered_vulnerabilities)
            
            # Ensure output directory exists
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            
            # Save filtered workbook
            self.workbook.save(output_path)
            
            # Get file size for logging
            file_size = os.path.getsize(output_path)
            file_size_mb = file_size / (1024 * 1024)
            
            logger.info(f"Filtered Excel export completed successfully:")
            logger.info(f"  - File: {output_path}")
            logger.info(f"  - Size: {file_size_mb:.2f} MB")
            logger.info(f"  - Sheets: Dependencies" + (", Vulnerabilities" if filtered_vulnerabilities else ""))
            
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to export filtered Excel file: {str(e)}")
            raise Exception(f"Filtered Excel export failed: {str(e)}")
            
        finally:
            # Restore original workbook
            self.workbook = original_workbook
    
    def export_policy_blocked(self, dependencies: List[ProcessedDependency], vulnerabilities: List[ProcessedVulnerability]) -> Optional[str]:
        """Export dependencies with LICENSE_POLICY_SETTING_BLOCK to separate Excel file."""
        if not dependencies:
            logger.info("No dependencies with LICENSE_POLICY_SETTING_BLOCK found, skipping blocked policy export")
            return None
            
        # Filter vulnerabilities to only those associated with policy blocked dependencies
        blocked_dep_names = {f"{dep.name}:{dep.version}" for dep in dependencies}
        filtered_vulnerabilities = [
            vuln for vuln in vulnerabilities 
            if f"{vuln.dependency_name}:{vuln.dependency_version}" in blocked_dep_names
        ]
        
        # Create new workbook for policy blocked export
        blocked_workbook = Workbook()
        blocked_workbook.remove(blocked_workbook.active)  # Remove default sheet
        original_workbook = self.workbook  # Store original workbook
        
        try:
            # Temporarily use blocked workbook
            self.workbook = blocked_workbook
            
            output_path = self._generate_policy_blocked_filename()
            logger.info(f"Starting LICENSE_POLICY_SETTING_BLOCK Excel export to {output_path}")
            logger.info(f"  - Policy blocked dependencies: {len(dependencies)}")
            logger.info(f"  - Associated vulnerabilities: {len(filtered_vulnerabilities)}")
            
            # Create sheets with policy blocked data (no license columns, no coloring)
            self._create_dependencies_sheet(dependencies, include_license_columns=False, apply_license_coloring=False)
            
            if filtered_vulnerabilities:
                self._create_vulnerabilities_sheet(filtered_vulnerabilities)
            
            # Ensure output directory exists
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            
            # Save blocked workbook
            self.workbook.save(output_path)
            
            # Get file size for logging
            file_size = os.path.getsize(output_path)
            file_size_mb = file_size / (1024 * 1024)
            
            logger.info(f"Policy blocked Excel export completed successfully:")
            logger.info(f"  - File: {output_path}")
            logger.info(f"  - Size: {file_size_mb:.2f} MB")
            logger.info(f"  - Sheets: Dependencies" + (", Vulnerabilities" if filtered_vulnerabilities else ""))
            
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to export policy blocked Excel file: {str(e)}")
            raise Exception(f"Policy blocked Excel export failed: {str(e)}")
            
        finally:
            # Restore original workbook
            self.workbook = original_workbook

    def export_policy_comment(self, dependencies: List[ProcessedDependency], vulnerabilities: List[ProcessedVulnerability]) -> Optional[str]:
        """Export dependencies with LICENSE_POLICY_SETTING_COMMENT to separate Excel file."""
        if not dependencies:
            logger.info("No dependencies with LICENSE_POLICY_SETTING_COMMENT found, skipping comment policy export")
            return None
            
        # Filter vulnerabilities to only those associated with policy comment dependencies
        comment_dep_names = {f"{dep.name}:{dep.version}" for dep in dependencies}
        filtered_vulnerabilities = [
            vuln for vuln in vulnerabilities 
            if f"{vuln.dependency_name}:{vuln.dependency_version}" in comment_dep_names
        ]
        
        # Create new workbook for policy comment export
        comment_workbook = Workbook()
        comment_workbook.remove(comment_workbook.active)  # Remove default sheet
        original_workbook = self.workbook  # Store original workbook
        
        try:
            # Temporarily use comment workbook
            self.workbook = comment_workbook
            
            output_path = self._generate_policy_comment_filename()
            logger.info(f"Starting LICENSE_POLICY_SETTING_COMMENT Excel export to {output_path}")
            logger.info(f"  - Policy comment dependencies: {len(dependencies)}")
            logger.info(f"  - Associated vulnerabilities: {len(filtered_vulnerabilities)}")
            
            # Create sheets with policy comment data (no license columns, no coloring)
            self._create_dependencies_sheet(dependencies, include_license_columns=False, apply_license_coloring=False)
            
            if filtered_vulnerabilities:
                self._create_vulnerabilities_sheet(filtered_vulnerabilities)
            
            # Ensure output directory exists
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            
            # Save comment workbook
            self.workbook.save(output_path)
            
            # Get file size for logging
            file_size = os.path.getsize(output_path)
            file_size_mb = file_size / (1024 * 1024)
            
            logger.info(f"Policy comment Excel export completed successfully:")
            logger.info(f"  - File: {output_path}")
            logger.info(f"  - Size: {file_size_mb:.2f} MB")
            logger.info(f"  - Sheets: Dependencies" + (", Vulnerabilities" if filtered_vulnerabilities else ""))
            
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to export policy comment Excel file: {str(e)}")
            raise Exception(f"Policy comment Excel export failed: {str(e)}")
            
        finally:
            # Restore original workbook
            self.workbook = original_workbook

    def export_ecosystem_pypi(self, dependencies: List[ProcessedDependency], vulnerabilities: List[ProcessedVulnerability]) -> Optional[str]:
        """Export PyPI ecosystem dependencies to separate Excel file."""
        if not dependencies:
            logger.info("No PyPI ecosystem dependencies found, skipping ecosystem export")
            return None
            
        # Filter vulnerabilities to only those associated with PyPI ecosystem dependencies
        pypi_dep_names = {f"{dep.name}:{dep.version}" for dep in dependencies}
        filtered_vulnerabilities = [
            vuln for vuln in vulnerabilities 
            if f"{vuln.dependency_name}:{vuln.dependency_version}" in pypi_dep_names
        ]
        
        # Create new workbook for ecosystem export
        ecosystem_workbook = Workbook()
        ecosystem_workbook.remove(ecosystem_workbook.active)  # Remove default sheet
        original_workbook = self.workbook  # Store original workbook
        
        try:
            # Temporarily use ecosystem workbook
            self.workbook = ecosystem_workbook
            
            output_path = self._generate_ecosystem_pypi_filename()
            logger.info(f"Starting PyPI ecosystem Excel export to {output_path}")
            logger.info(f"  - PyPI ecosystem dependencies: {len(dependencies)}")
            logger.info(f"  - Associated vulnerabilities: {len(filtered_vulnerabilities)}")
            
            # Create sheets with ecosystem data (no license columns, no coloring)
            self._create_dependencies_sheet(dependencies, include_license_columns=False, apply_license_coloring=False)
            
            if filtered_vulnerabilities:
                self._create_vulnerabilities_sheet(filtered_vulnerabilities)
            
            # Ensure output directory exists
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            
            # Save ecosystem workbook
            self.workbook.save(output_path)
            
            # Get file size for logging
            file_size = os.path.getsize(output_path)
            file_size_mb = file_size / (1024 * 1024)
            
            logger.info(f"PyPI ecosystem Excel export completed successfully:")
            logger.info(f"  - File: {output_path}")
            logger.info(f"  - Size: {file_size_mb:.2f} MB")
            logger.info(f"  - Sheets: Dependencies" + (", Vulnerabilities" if filtered_vulnerabilities else ""))
            
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to export PyPI ecosystem Excel file: {str(e)}")
            raise Exception(f"PyPI ecosystem Excel export failed: {str(e)}")
            
        finally:
            # Restore original workbook
            self.workbook = original_workbook
    
    def __del__(self):
        """Clean up workbook resources."""
        if hasattr(self, 'workbook'):
            try:
                self.workbook.close()
            except:
                pass