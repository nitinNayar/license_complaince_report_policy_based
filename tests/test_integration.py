"""
Integration tests for the Semgrep Dependencies Export Tool.

Tests the complete end-to-end workflow with mocked API responses.
"""

import json
import os
import pytest
import responses
import sys
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

# Add src to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from semgrep_deps_export.config import Config
from semgrep_deps_export.main import SemgrepDepsExporter
from semgrep_deps_export.api_client import SemgrepAPIClient


class TestEndToEndIntegration:
    """Integration tests for complete workflow."""
    
    @pytest.fixture
    def config(self):
        """Create test configuration."""
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "test_output.xlsx")
            yield Config(
                token="test_token_12345678901234567890",
                deployment_id="test_deployment_123",
                output_path=output_path,
                log_level="INFO"
            )
    
    @pytest.fixture
    def sample_api_responses(self):
        """Create sample API responses for testing."""
        return [
            {
                "dependencies": [
                    {
                        "id": "dep-1",
                        "name": "lodash",
                        "version": "4.17.21",
                        "ecosystem": "npm",
                        "package_manager": "npm",
                        "licenses": ["MIT"],
                        "vulnerabilities": [
                            {
                                "id": "GHSA-35jh-r3h4-6jhm",
                                "severity": "high",
                                "description": "Command injection vulnerability"
                            }
                        ],
                        "first_seen": "2023-01-01T10:00:00Z",
                        "last_seen": "2023-12-01T15:30:00Z",
                        "projects": ["web-app", "mobile-app"]
                    },
                    {
                        "id": "dep-2",
                        "name": "express",
                        "version": "4.18.2",
                        "ecosystem": "npm",
                        "package_manager": "npm",
                        "licenses": ["MIT"],
                        "vulnerabilities": [],
                        "first_seen": "2023-02-01T10:00:00Z",
                        "last_seen": "2023-12-01T15:30:00Z",
                        "projects": ["web-app"]
                    }
                ],
                "cursor": "page2_cursor",
                "has_more": True
            },
            {
                "dependencies": [
                    {
                        "id": "dep-3",
                        "name": "react",
                        "version": "18.2.0",
                        "ecosystem": "npm",
                        "package_manager": "npm",
                        "licenses": ["MIT"],
                        "vulnerabilities": [
                            {
                                "id": "CVE-2023-1234",
                                "severity": "medium",
                                "description": "Medium severity issue"
                            },
                            {
                                "id": "CVE-2023-5678",
                                "severity": "critical",
                                "description": "Critical security vulnerability"
                            }
                        ],
                        "first_seen": "2023-03-01T10:00:00Z",
                        "last_seen": "2023-12-01T15:30:00Z",
                        "projects": ["web-app", "admin-panel"]
                    }
                ],
                "has_more": False
            }
        ]
    
    @responses.activate
    def test_full_export_workflow(self, config, sample_api_responses):
        """Test the complete export workflow."""
        base_url = SemgrepAPIClient.BASE_URL
        endpoint_url = f"{base_url}/deployments/{config.deployment_id}/dependencies"
        
        # Mock first API call
        responses.add(
            responses.POST,
            endpoint_url,
            json=sample_api_responses[0],
            status=200
        )
        
        # Mock second API call
        responses.add(
            responses.POST,
            endpoint_url,
            json=sample_api_responses[1],
            status=200
        )
        
        # Create and run exporter
        exporter = SemgrepDepsExporter(config)
        
        with patch('semgrep_deps_export.utils.setup_logging'):
            success = exporter.run()
        
        # Verify success
        assert success is True
        
        # Verify output file was created
        assert os.path.exists(config.output_path)
        assert os.path.getsize(config.output_path) > 0
        
        # Verify API calls were made correctly
        assert len(responses.calls) == 2
        
        # Check first request had no cursor
        first_request = json.loads(responses.calls[0].request.body)
        assert "cursor" not in first_request or first_request["cursor"] is None
        
        # Check second request had cursor from first response
        second_request = json.loads(responses.calls[1].request.body)
        assert second_request["cursor"] == "page2_cursor"
    
    @responses.activate
    def test_api_authentication_failure(self, config, sample_api_responses):
        """Test handling of authentication failure."""
        endpoint_url = f"{SemgrepAPIClient.BASE_URL}/deployments/{config.deployment_id}/dependencies"
        
        # Mock authentication failure
        responses.add(
            responses.POST,
            endpoint_url,
            json={"message": "Invalid token"},
            status=401
        )
        
        exporter = SemgrepDepsExporter(config)
        
        with patch('semgrep_deps_export.utils.setup_logging'):
            success = exporter.run()
        
        assert success is False
    
    @responses.activate
    def test_rate_limiting_with_retry(self, config, sample_api_responses):
        """Test rate limiting handling with retry logic."""
        endpoint_url = f"{SemgrepAPIClient.BASE_URL}/deployments/{config.deployment_id}/dependencies"
        
        # Mock rate limit error first, then success
        responses.add(
            responses.POST,
            endpoint_url,
            json={"message": "Rate limit exceeded"},
            status=429
        )
        
        responses.add(
            responses.POST,
            endpoint_url,
            json=sample_api_responses[0],
            status=200
        )
        
        responses.add(
            responses.POST,
            endpoint_url,
            json=sample_api_responses[1],
            status=200
        )
        
        exporter = SemgrepDepsExporter(config)
        
        # Mock time.sleep to avoid actual delays in tests
        with patch('time.sleep'), patch('semgrep_deps_export.utils.setup_logging'):
            success = exporter.run()
        
        assert success is True
        # Should have made 3 total calls (1 failed + 2 successful)
        assert len(responses.calls) == 3
    
    @responses.activate
    def test_empty_dependencies_response(self, config):
        """Test handling of empty dependencies response."""
        endpoint_url = f"{SemgrepAPIClient.BASE_URL}/deployments/{config.deployment_id}/dependencies"
        
        # Mock empty response
        responses.add(
            responses.POST,
            endpoint_url,
            json={"dependencies": [], "has_more": False},
            status=200
        )
        
        exporter = SemgrepDepsExporter(config)
        
        with patch('semgrep_deps_export.utils.setup_logging'):
            success = exporter.run()
        
        assert success is False  # Should fail with no dependencies
    
    @responses.activate 
    def test_data_processing_and_excel_generation(self, config, sample_api_responses):
        """Test that data is properly processed and Excel file contains expected data."""
        endpoint_url = f"{SemgrepAPIClient.BASE_URL}/deployments/{config.deployment_id}/dependencies"
        
        # Mock API responses
        for response in sample_api_responses:
            responses.add(
                responses.POST,
                endpoint_url,
                json=response,
                status=200
            )
        
        exporter = SemgrepDepsExporter(config)
        
        with patch('semgrep_deps_export.utils.setup_logging'):
            success = exporter.run()
        
        assert success is True
        
        # Verify file exists and has reasonable size
        assert os.path.exists(config.output_path)
        file_size = os.path.getsize(config.output_path)
        assert file_size > 5000  # Should be reasonably large Excel file
        
        # Verify processing stats
        summary = exporter.data_processor.get_processing_summary()
        assert summary["dependencies"]["total"] == 3
        assert summary["dependencies"]["with_vulnerabilities"] == 2
        assert summary["dependencies"]["without_vulnerabilities"] == 1
        assert summary["vulnerabilities"]["total"] == 3
        assert summary["vulnerabilities"]["critical"] == 1
        assert summary["vulnerabilities"]["high"] == 1
        assert summary["vulnerabilities"]["medium"] == 1
    
    @responses.activate
    def test_network_error_handling(self, config):
        """Test handling of network errors."""
        endpoint_url = f"{SemgrepAPIClient.BASE_URL}/deployments/{config.deployment_id}/dependencies"
        
        # Mock network error
        responses.add(
            responses.POST,
            endpoint_url,
            body=Exception("Network error")
        )
        
        exporter = SemgrepDepsExporter(config)
        
        with patch('semgrep_deps_export.utils.setup_logging'):
            success = exporter.run()
        
        assert success is False
    
    def test_configuration_validation(self):
        """Test configuration validation."""
        # Test missing token
        with pytest.raises(ValueError, match="SEMGREP_APP_TOKEN is required"):
            Config(token="", deployment_id="test_deployment")
        
        # Test missing deployment_id
        with pytest.raises(ValueError, match="deployment_id is required"):
            Config(token="test_token", deployment_id="")
    
    @responses.activate
    def test_excel_file_structure(self, config, sample_api_responses):
        """Test that Excel file has correct structure."""
        endpoint_url = f"{SemgrepAPIClient.BASE_URL}/deployments/{config.deployment_id}/dependencies"
        
        # Mock API responses
        for response in sample_api_responses:
            responses.add(
                responses.POST,
                endpoint_url,
                json=response,
                status=200
            )
        
        exporter = SemgrepDepsExporter(config)
        
        with patch('semgrep_deps_export.utils.setup_logging'):
            success = exporter.run()
        
        assert success is True
        
        # Try to read the Excel file to verify structure
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(config.output_path)
            
            # Check worksheets exist
            assert "Summary" in wb.sheetnames
            assert "Dependencies" in wb.sheetnames
            assert "Vulnerabilities" in wb.sheetnames
            
            # Check Dependencies sheet has data
            deps_sheet = wb["Dependencies"]
            assert deps_sheet.max_row > 1  # Has header + data rows
            
            # Check Vulnerabilities sheet has data
            vulns_sheet = wb["Vulnerabilities"]
            assert vulns_sheet.max_row > 1  # Has header + data rows
            
            wb.close()
            
        except ImportError:
            # If openpyxl not available in test environment, skip validation
            pass
    
    @patch.dict(os.environ, {
        'SEMGREP_APP_TOKEN': 'env_test_token_123456789012',
        'SEMGREP_DEPLOYMENT_ID': 'env_test_deployment'
    })
    @patch('sys.argv', ['script.py'])
    @responses.activate
    def test_environment_variable_config(self, sample_api_responses):
        """Test loading configuration from environment variables."""
        from semgrep_deps_export.config import ConfigManager
        
        config_manager = ConfigManager()
        config = config_manager.load_config()
        
        assert config.token == "env_test_token_123456789012"
        assert config.deployment_id == "env_test_deployment"
        
        # Test the export process with env config
        endpoint_url = f"{SemgrepAPIClient.BASE_URL}/deployments/{config.deployment_id}/dependencies"
        
        responses.add(
            responses.POST,
            endpoint_url,
            json=sample_api_responses[0],
            status=200
        )
        
        responses.add(
            responses.POST,
            endpoint_url,
            json=sample_api_responses[1],
            status=200
        )
        
        exporter = SemgrepDepsExporter(config)
        
        with patch('semgrep_deps_export.utils.setup_logging'):
            success = exporter.run()
        
        assert success is True